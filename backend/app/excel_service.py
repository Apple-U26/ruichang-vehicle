"""Excel 导入导出服务。

支持车辆台账导入以及包含车辆、里程、维保、报销、月度汇总的完整导出。
"""

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import (
    FuelRecord,
    MileageRecord,
    MaintenanceRecord,
    Project,
    Reimbursement,
    ReimbursementDetail,
    User,
    Vehicle,
    ViolationRecord,
    Welder,
    WelderInspection,
)


STATUS_LABELS = {
    "ACTIVE": "启用",
    "MAINTENANCE": "维保中",
    "DISABLED": "已停用",
    "RETURNED": "已归还",
    "OUT": "出车中",
    "CLOSED": "已收车",
    "DRAFT": "草稿",
    "SUBMITTED": "待项目审核",
    "PROJECT_APPROVED": "待财务审核",
    "APPROVED": "已通过",
    "REJECTED": "已退回",
    "UNPROCESSED": "未处理",
    "PROCESSED": "已处理",
}

OWNERSHIP_LABELS = {
    "COMPANY": "公司车辆",
    "RENTAL": "租赁车辆",
    "TEMPORARY": "临时车辆",
    "OTHER": "其他",
}

MAINTENANCE_TYPE_LABELS = {
    "MAINTENANCE": "保养",
    "REPAIR": "维修",
    "INSPECTION": "年检",
    "INSURANCE": "保险",
    "MONTHLY": "月检",
    "WEEKLY": "周检",
    "DAILY": "日检",
}


class ExcelService:
    HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
    HEADER_FILL = PatternFill("solid", fgColor="2F6FAD")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    CELL_ALIGN = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _write_table_to_sheet(
        ws: Worksheet,
        data: List[Dict[str, Any]],
        headers: List[str],
        start_row: int = 1,
    ) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = ExcelService.HEADER_FONT
            cell.fill = ExcelService.HEADER_FILL
            cell.alignment = ExcelService.HEADER_ALIGN
            cell.border = ExcelService.BORDER

        for row_idx, row_data in enumerate(
            data or [{}], start=start_row + 1
        ):
            for col_idx, key in enumerate(headers, start=1):
                value = row_data.get(key, "")
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = ExcelService.CELL_ALIGN
                cell.border = ExcelService.BORDER

        if not data:
            ws.cell(
                row=start_row + 1,
                column=1,
                value="暂无数据",
            ).alignment = ExcelService.CELL_ALIGN

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1]))
            for row in range(
                start_row + 1,
                start_row + max(len(data), 1) + 1,
            ):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(
                max_len + 3, 40
            )

    @staticmethod
    def _get_vehicles_data(db: Session) -> List[Dict[str, Any]]:
        vehicles = db.scalars(select(Vehicle).order_by(Vehicle.id)).all()
        result = []
        for v in vehicles:
            result.append(
                {
                    "车辆编码": v.vehicle_code,
                    "车牌号": v.plate_no,
                    "所属项目": v.project.name if v.project else "",
                    "项目负责人": v.project_manager or "",
                    "车管员": v.vehicle_manager or "",
                    "所有权": OWNERSHIP_LABELS.get(
                        v.ownership, v.ownership
                    ),
                    "初始里程(km)": float(v.initial_mileage),
                    "当前里程(km)": float(v.current_mileage),
                    "状态": STATUS_LABELS.get(v.status, v.status),
                    "车龄(年)": v.vehicle_age or "",
                    "违章信息": v.violation_info or "",
                    "备注": v.remark or "",
                }
            )
        return result

    @staticmethod
    def _get_mileages_data(db: Session) -> List[Dict[str, Any]]:
        stmt = (
            select(MileageRecord, Vehicle.plate_no)
            .join(Vehicle, Vehicle.id == MileageRecord.vehicle_id)
            .order_by(MileageRecord.id)
        )
        rows = db.execute(stmt).all()
        result = []
        for record, plate_no in rows:
            result.append(
                {
                    "车牌号": plate_no,
                    "日期": record.trip_date,
                    "出车里程(km)": float(record.out_mileage),
                    "收车里程(km)": (
                        float(record.in_mileage)
                        if record.in_mileage is not None
                        else ""
                    ),
                    "行驶里程(km)": float(record.distance),
                    "驾驶员": record.driver_name or "",
                    "出发地": record.departure or "",
                    "目的地": record.destination or "",
                    "事由": record.purpose or "",
                    "状态": STATUS_LABELS.get(record.status, record.status),
                    "异常": "是" if record.abnormal else "否",
                    "异常原因": record.abnormal_reason or "",
                }
            )
        return result

    @staticmethod
    def _get_maintenances_data(db: Session) -> List[Dict[str, Any]]:
        stmt = (
            select(MaintenanceRecord, Vehicle.plate_no)
            .join(
                Vehicle,
                Vehicle.id == MaintenanceRecord.vehicle_id,
            )
            .order_by(MaintenanceRecord.id)
        )
        rows = db.execute(stmt).all()
        result = []
        for record, plate_no in rows:
            result.append(
                {
                    "车牌号": plate_no,
                    "维保日期": record.maintenance_date,
                    "当前里程(km)": float(record.current_mileage),
                    "类型": MAINTENANCE_TYPE_LABELS.get(
                        record.maintenance_type, record.maintenance_type
                    ),
                    "项目": record.items,
                    "金额(元)": float(record.amount),
                    "服务商": record.service_provider or "",
                    "操作人": record.operator_name or "",
                    "下次里程(km)": (
                        float(record.next_mileage)
                        if record.next_mileage is not None
                        else ""
                    ),
                    "下次日期": record.next_date or "",
                    "备注": record.remark or "",
                }
            )
        return result

    @staticmethod
    def _get_reimbursements_data(db: Session) -> List[Dict[str, Any]]:
        stmt = (
            select(Reimbursement, Vehicle.plate_no)
            .join(Vehicle, Vehicle.id == Reimbursement.vehicle_id)
            .order_by(Reimbursement.id)
        )
        rows = db.execute(stmt).all()
        result = []
        for r, plate_no in rows:
            result.append(
                {
                    "报销单号": r.reimbursement_no,
                    "报销月份": r.reimbursement_month,
                    "车牌号": plate_no,
                    "申请人": r.applicant_name,
                    "总金额(元)": float(r.total_amount),
                    "状态": STATUS_LABELS.get(r.status, r.status),
                    "退回原因": r.reject_reason or "",
                    "创建时间": r.created_at,
                }
            )
        return result

    @staticmethod
    def _get_violations_data(db: Session) -> List[Dict[str, Any]]:
        stmt = (
            select(ViolationRecord, Vehicle.plate_no)
            .join(Vehicle, Vehicle.id == ViolationRecord.vehicle_id)
            .order_by(ViolationRecord.id)
        )
        rows = db.execute(stmt).all()
        result = []
        for record, plate_no in rows:
            result.append(
                {
                    "车牌号": plate_no,
                    "违章日期": record.violation_date,
                    "违章类型": record.violation_type or "",
                    "违章地点": record.location or "",
                    "扣分": record.points or "",
                    "罚款金额(元)": float(record.fine_amount),
                    "状态": STATUS_LABELS.get(
                        record.status, record.status
                    ),
                    "附件": record.attachment_url or "",
                    "处理人": record.handler_name or "",
                    "备注": record.remark or "",
                }
            )
        return result

    @staticmethod
    def _get_fuels_data(db: Session) -> List[Dict[str, Any]]:
        stmt = (
            select(FuelRecord, Vehicle.plate_no)
            .join(Vehicle, Vehicle.id == FuelRecord.vehicle_id)
            .order_by(FuelRecord.id)
        )
        rows = db.execute(stmt).all()
        result = []
        for record, plate_no in rows:
            result.append(
                {
                    "车牌号": plate_no,
                    "加油日期": record.fuel_date,
                    "加油量(升)": float(record.liters),
                    "单价(元/升)": float(record.unit_price),
                    "金额(元)": float(record.total_amount),
                    "里程(km)": (
                        float(record.mileage)
                        if record.mileage is not None
                        else ""
                    ),
                    "加油站": record.station or "",
                    "发票号": record.invoice_no or "",
                    "备注": record.remark or "",
                }
            )
        return result

    @staticmethod
    def _get_welders_data(db: Session) -> List[Dict[str, Any]]:
        welders = db.scalars(
            select(Welder).order_by(Welder.id)
        ).all()
        result = []
        for welder in welders:
            result.append(
                {
                    "焊机编码": welder.welder_code,
                    "焊机编号": welder.welder_no,
                    "所在地": welder.location or "",
                    "所属项目": (
                        welder.project.name if welder.project else ""
                    ),
                    "项目负责人": (
                        welder.project.manager_name
                        if welder.project
                        else ""
                    ),
                    "焊机负责人": welder.welder_manager or "",
                    "状态": STATUS_LABELS.get(
                        welder.status, welder.status
                    ),
                    "备注": welder.remark or "",
                }
            )
        return result

    @staticmethod
    def _get_welder_inspections_data(
        db: Session,
    ) -> List[Dict[str, Any]]:
        stmt = (
            select(WelderInspection)
            .join(
                Welder,
                Welder.id == WelderInspection.welder_id,
            )
            .order_by(
                Welder.welder_no,
                WelderInspection.inspection_date,
            )
        )
        rows = db.scalars(stmt).all()
        result = []
        for row in rows:
            result.append(
                {
                    "焊机编号": (
                        row.welder.welder_no if row.welder else ""
                    ),
                    "所在地": row.location or "",
                    "项目": (
                        row.welder.project.name
                        if row.welder and row.welder.project
                        else ""
                    ),
                    "项目负责人": (
                        row.welder.project.manager_name
                        if row.welder and row.welder.project
                        else ""
                    ),
                    "焊机负责人": (
                        row.welder.welder_manager if row.welder else ""
                    ),
                    "日期": row.inspection_date,
                    "巡检类型": MAINTENANCE_TYPE_LABELS.get(
                        row.inspection_type, row.inspection_type
                    ),
                    "是否完成": "是" if row.completed else "否",
                    "附件": row.attachment_url or "",
                    "操作人": row.operator_name or "",
                    "设备状态": (
                        "正常"
                        if row.device_status == "NORMAL"
                        else "故障"
                    ),
                    "备注": row.remark or "",
                    "维修说明": row.repair_note or "",
                }
            )
        return result

    @staticmethod
    def _get_dashboard_data(db: Session) -> Dict[str, Any]:
        total_vehicles = (
            db.scalar(select(func.count(Vehicle.id))) or 0
        )
        active_vehicles = (
            db.scalar(
                select(func.count(Vehicle.id)).where(
                    Vehicle.status == "ACTIVE"
                )
            )
            or 0
        )
        total_mileage = (
            db.scalar(
                select(
                    func.coalesce(func.sum(MileageRecord.distance), 0)
                ).where(MileageRecord.status == "CLOSED")
            )
            or 0
        )
        total_reimbursement = (
            db.scalar(
                select(
                    func.coalesce(
                        func.sum(Reimbursement.total_amount), 0
                    )
                ).where(Reimbursement.status == "APPROVED")
            )
            or 0
        )
        pending_count = (
            db.scalar(
                select(func.count(Reimbursement.id)).where(
                    Reimbursement.status.in_(
                        ["SUBMITTED", "PROJECT_APPROVED"]
                    )
                )
            )
            or 0
        )

        overview = {
            "总车辆数": total_vehicles,
            "活跃车辆数": active_vehicles,
            "累计总里程(km)": float(total_mileage),
            "累计报销总额(元)": float(total_reimbursement),
            "待审批报销单数": pending_count,
        }

        mileage_monthly = db.execute(
            select(
                func.date_format(
                    MileageRecord.trip_date, "%Y-%m"
                ).label("month"),
                func.coalesce(
                    func.sum(MileageRecord.distance), 0
                ).label("mileage"),
            )
            .where(MileageRecord.status == "CLOSED")
            .group_by("month")
            .order_by("month")
        ).all()

        reimb_monthly = db.execute(
            select(
                Reimbursement.reimbursement_month.label("month"),
                func.coalesce(
                    func.sum(Reimbursement.total_amount), 0
                ).label("total"),
            )
            .where(Reimbursement.status == "APPROVED")
            .group_by(Reimbursement.reimbursement_month)
            .order_by(Reimbursement.reimbursement_month)
        ).all()

        expense_monthly = db.execute(
            select(
                Reimbursement.reimbursement_month.label("month"),
                ReimbursementDetail.expense_type.label("type"),
                func.coalesce(
                    func.sum(ReimbursementDetail.amount), 0
                ).label("amount"),
            )
            .join(
                Reimbursement,
                Reimbursement.id
                == ReimbursementDetail.reimbursement_id,
            )
            .where(Reimbursement.status == "APPROVED")
            .group_by(
                Reimbursement.reimbursement_month,
                ReimbursementDetail.expense_type,
            )
        ).all()

        pending_monthly = db.execute(
            select(
                Reimbursement.reimbursement_month.label("month"),
                func.count(Reimbursement.id).label("pending"),
            )
            .where(
                Reimbursement.status.in_(
                    ["SUBMITTED", "PROJECT_APPROVED"]
                )
            )
            .group_by(Reimbursement.reimbursement_month)
        ).all()

        monthly_map: Dict[str, Dict[str, Any]] = {}

        def ensure_month(month: str) -> Dict[str, Any]:
            if month not in monthly_map:
                monthly_map[month] = {
                    "month": month,
                    "总里程(km)": 0,
                    "报销总额(元)": 0,
                    "燃油费(元)": 0,
                    "维保费(元)": 0,
                    "路桥费(元)": 0,
                    "停车费(元)": 0,
                    "其他费(元)": 0,
                    "里程补助费(元)": 0,
                    "待审批数": 0,
                }
            return monthly_map[month]

        for month, mileage in mileage_monthly:
            ensure_month(month)["总里程(km)"] = float(mileage)

        for month, total in reimb_monthly:
            ensure_month(month)["报销总额(元)"] = float(total)

        expense_keys = {
            "FUEL": "燃油费(元)",
            "MAINTENANCE": "维保费(元)",
            "TOLL": "路桥费(元)",
            "PARKING": "停车费(元)",
            "OTHER": "其他费(元)",
            "MILEAGE_ALLOWANCE": "里程补助费(元)",
        }
        for month, expense_type, amount in expense_monthly:
            key = expense_keys.get(expense_type)
            if key:
                ensure_month(month)[key] = float(amount)

        for month, pending in pending_monthly:
            ensure_month(month)["待审批数"] = int(pending)

        monthly_details = sorted(
            monthly_map.values(), key=lambda x: x["month"]
        )

        return {
            "overview": overview,
            "monthly": monthly_details,
        }

    @staticmethod
    def export_full_workbook(db: Session) -> io.BytesIO:
        wb = Workbook()

        sections = [
            ("车辆信息", ExcelService._get_vehicles_data),
            ("里程记录", ExcelService._get_mileages_data),
            ("维保记录", ExcelService._get_maintenances_data),
            ("违章记录", ExcelService._get_violations_data),
            ("油费记录", ExcelService._get_fuels_data),
            ("焊机档案", ExcelService._get_welders_data),
            ("焊机巡检", ExcelService._get_welder_inspections_data),
            ("报销单", ExcelService._get_reimbursements_data),
        ]

        for index, (title, getter) in enumerate(sections):
            ws = wb.active if index == 0 else wb.create_sheet(title)
            ws.title = title
            data = getter(db)
            headers = list(data[0].keys()) if data else ["无数据"]
            ExcelService._write_table_to_sheet(
                ws, data, headers, start_row=1
            )

        ws5 = wb.create_sheet("月度汇总仪表盘")
        dashboard = ExcelService._get_dashboard_data(db)
        overview = dashboard["overview"]
        monthly = dashboard["monthly"]

        row = 1
        ws5.cell(row=row, column=1, value="总览")
        ws5.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        for idx, (key, value) in enumerate(overview.items()):
            ws5.cell(row=row + idx, column=1, value=key)
            ws5.cell(row=row + idx, column=2, value=value)
            for col in (1, 2):
                cell = ws5.cell(row=row + idx, column=col)
                cell.border = ExcelService.BORDER
                cell.alignment = ExcelService.CELL_ALIGN
        ws5.column_dimensions["A"].width = 20
        ws5.column_dimensions["B"].width = 24
        row += len(overview) + 2

        ws5.cell(row=row, column=1, value="月度明细")
        ws5.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        if monthly:
            headers = list(monthly[0].keys())
            ExcelService._write_table_to_sheet(
                ws5, monthly, headers, start_row=row
            )
        else:
            ws5.cell(row=row, column=1, value="暂无月度数据")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def _parse_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except Exception:
        return Decimal("0")


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_status(value: Any) -> str:
    text = str(value or "").strip().upper()
    status_map = {
        "启用": "ACTIVE",
        "ACTIVE": "ACTIVE",
        "维保中": "MAINTENANCE",
        "保养中": "MAINTENANCE",
        "MAINTENANCE": "MAINTENANCE",
        "已停用": "DISABLED",
        "停用": "DISABLED",
        "DISABLED": "DISABLED",
        "已归还": "RETURNED",
        "RETURNED": "RETURNED",
    }
    return status_map.get(text, "ACTIVE")


def _parse_ownership(value: Any) -> str:
    text = str(value or "").strip()
    if "租赁" in text:
        return "RENTAL"
    if "公司" in text:
        return "COMPANY"
    if "临时" in text:
        return "TEMPORARY"
    upper_text = text.upper()
    if "COMPANY" in upper_text:
        return "COMPANY"
    if "RENTAL" in upper_text:
        return "RENTAL"
    if "TEMPORARY" in upper_text:
        return "TEMPORARY"
    return "OTHER"


def _parse_maintenance_type(value: Any) -> str:
    text = str(value or "").strip()
    if "维修" in text:
        return "REPAIR"
    if "年检" in text:
        return "INSPECTION"
    if "保险" in text:
        return "INSURANCE"
    return "MAINTENANCE"


def _parse_month(value: Any) -> tuple[int, int] | None:
    text = str(value or "").strip()
    if not text:
        return None

    import re

    match = re.search(r"(20\d{2})[年\-/.](\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month

    clean = text.replace("月", "").strip()
    if clean.isdigit():
        month = int(clean)
        if 1 <= month <= 12:
            now = datetime.now()
            year = now.year
            if month > now.month:
                year -= 1
            return year, month
    return None


def _parse_date(value: Any, default: date | None = None) -> date | None:
    if value is None or value == "":
        return default
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return default


FIELD_ALIASES = {
    "plate_no": ["车牌号"],
    "project_name": ["所属项目", "项目"],
    "project_manager": ["项目负责人", "项目经理"],
    "vehicle_manager": ["车管员", "车辆负责人"],
    "ownership": ["所有权", "车辆归属"],
    "initial_mileage": ["初始里程(km)", "初始里程"],
    "current_mileage": ["当前里程(km)", "当前里程"],
    "appearance_url": ["车辆外观"],
    "vehicle_age": ["车辆年限", "车龄(年)", "车龄"],
    "violation_info": ["违章信息", "违章情况"],
    "status": ["状态"],
    "remark": ["备注"],
    "month": ["日期（月份）", "日期(月份)", "报销月份", "月份"],
    "out_mileage": ["出车里程(km)", "出车里程"],
    "in_mileage": ["收车里程(km)", "收车里程"],
    "distance": ["本次里程(km)", "本次里程", "行驶里程(km)", "行驶里程"],
    "driver_name": ["驾驶员", "驾驶人", "车辆负责人"],
    "maintenance_date": ["保养日期", "维保日期", "日期"],
    "maintenance_type": ["保养类型", "维保类型", "类型"],
    "items": ["保养项目", "维保项目", "项目"],
    "amount": ["花费金额(元)", "金额(元)", "金额"],
    "service_provider": ["维修方", "服务商"],
    "operator_name": ["车辆负责人", "操作人"],
    "next_mileage": ["下次保养里程(km)", "下次里程(km)", "下次里程"],
    "related_mileage": ["对应行程里程(km)", "对应行程里程"],
    "fuel": ["油费(元)", "油费"],
    "maintenance": ["维保费用(元)", "维保费用"],
    "toll": ["路桥停车费(元)", "路桥停车费", "路桥费(元)", "路桥费"],
    "parking": ["停车费(元)", "停车费"],
    "other": ["其他费用(元)", "其他费用", "其他费(元)", "其他费"],
    "total": ["报销总额(元)", "报销总额"],
    "applicant": ["报销人", "申请人"],
    "reimbursement_status": ["审核状态"],
    "welder_code": ["焊机编码"],
    "welder_no": ["焊机编号"],
    "welder_location": ["所在地", "所在地（市级行政区）", "所在地(市级行政区）"],
    "welder_manager": ["焊机负责人"],
    "welder_status": ["状态"],
}


def _normalize_header(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .replace("（", "(")
        .replace("）", ")")
        .replace(" ", "")
    )


def _find_sheet(workbook, keywords: list[str]):
    for name in workbook.sheetnames:
        if any(keyword in name for keyword in keywords):
            return workbook[name]
    return None


def _locate_header(ws, required_key="车牌号"):
    max_scan = min(ws.max_row or 1, 30)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True),
        1,
    ):
        headers = {}
        for col_idx, value in enumerate(row):
            if value is None or str(value).strip() == "":
                continue
            headers[_normalize_header(value)] = col_idx
        if required_key in headers:
            return row_idx, headers
    return None, {}


def _cell(row, headers: Dict[str, int], field: str) -> Any:
    for alias in FIELD_ALIASES.get(field, []):
        index = headers.get(alias)
        if index is not None and index < len(row):
            value = row[index]
            if value is not None and str(value).strip() != "":
                return value
    return None


def _get_or_create_project(
    db: Session, name: str, manager_name: str | None
) -> Project | None:
    name = str(name or "").strip()
    if not name:
        return None
    project = db.scalar(select(Project).where(Project.name == name))
    if project is None:
        project = Project(
            name=name,
            manager_name=manager_name or None,
            enabled=True,
        )
        db.add(project)
        db.flush()
    elif manager_name and not project.manager_name:
        project.manager_name = manager_name
    return project


def _next_vehicle_code(db: Session) -> str:
    max_id = db.scalar(select(func.max(Vehicle.id))) or 0
    while True:
        max_id += 1
        code = f"CL-{max_id:06d}"
        exists = db.scalar(
            select(Vehicle).where(Vehicle.vehicle_code == code)
        )
        if not exists:
            return code


def _next_welder_code(db: Session) -> str:
    max_id = db.scalar(select(func.max(Welder.id))) or 0
    while True:
        max_id += 1
        code = f"HJ-{max_id:06d}"
        exists = db.scalar(
            select(Welder).where(Welder.welder_code == code)
        )
        if not exists:
            return code


def _next_reimbursement_no(db: Session, month: str) -> str:
    sequence = (
        db.scalar(
            select(func.count(Reimbursement.id)).where(
                Reimbursement.reimbursement_month == month
            )
        )
        or 0
    )
    return f"BX-{month.replace('-', '')}-{sequence + 1:04d}"


def _import_vehicles(
    db: Session, workbook, user: User
) -> tuple[int, int, int]:
    ws = _find_sheet(workbook, ["车辆基础信息", "车辆信息"])
    if ws is None:
        ws = workbook.worksheets[0]

    header_row, headers = _locate_header(ws)
    if "车牌号" not in headers:
        return 0, 0, 0

    created = 0
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        plate_no = str(_cell(row, headers, "plate_no") or "").strip().upper()
        if not plate_no:
            skipped += 1
            continue

        project_name = str(_cell(row, headers, "project_name") or "").strip()
        project_manager = str(
            _cell(row, headers, "project_manager") or ""
        ).strip()
        vehicle_manager = str(
            _cell(row, headers, "vehicle_manager") or ""
        ).strip()
        initial_mileage = _parse_decimal(
            _cell(row, headers, "initial_mileage")
        )
        current_mileage = _parse_decimal(
            _cell(row, headers, "current_mileage")
        )
        vehicle_age = _parse_int(_cell(row, headers, "vehicle_age"))
        appearance_url = str(
            _cell(row, headers, "appearance_url") or ""
        ).strip()
        violation_info = str(
            _cell(row, headers, "violation_info") or ""
        ).strip()
        remark = str(_cell(row, headers, "remark") or "").strip()
        ownership = _parse_ownership(_cell(row, headers, "ownership"))
        status = _parse_status(_cell(row, headers, "status"))

        project = _get_or_create_project(
            db, project_name, project_manager
        )
        vehicle = db.scalar(
            select(Vehicle).where(Vehicle.plate_no == plate_no)
        )

        if vehicle is None:
            vehicle = Vehicle(
                vehicle_code=_next_vehicle_code(db),
                plate_no=plate_no,
                project_id=project.id if project else None,
                project_manager=project_manager or None,
                vehicle_manager=vehicle_manager or None,
                ownership=ownership,
                initial_mileage=initial_mileage,
                current_mileage=current_mileage or initial_mileage,
                status=status,
                vehicle_age=vehicle_age,
                violation_info=violation_info or None,
                appearance_url=appearance_url or None,
                remark=remark or None,
            )
            db.add(vehicle)
            db.flush()
            created += 1
        else:
            if project:
                vehicle.project_id = project.id
            if project_manager:
                vehicle.project_manager = project_manager
            if vehicle_manager:
                vehicle.vehicle_manager = vehicle_manager
            if initial_mileage:
                vehicle.initial_mileage = initial_mileage
            if (
                current_mileage
                and current_mileage > vehicle.current_mileage
            ):
                vehicle.current_mileage = current_mileage
            if vehicle_age is not None:
                vehicle.vehicle_age = vehicle_age
            if appearance_url:
                vehicle.appearance_url = appearance_url
            if violation_info:
                vehicle.violation_info = violation_info
            if remark:
                vehicle.remark = remark
            vehicle.ownership = ownership
            vehicle.status = status
            updated += 1

    return created, updated, skipped


def _import_mileages(
    db: Session, workbook, user: User
) -> int:
    ws = _find_sheet(workbook, ["里程"])
    if ws is None:
        return 0
    header_row, headers = _locate_header(ws)
    if "车牌号" not in headers:
        return 0

    created = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        plate_no = str(_cell(row, headers, "plate_no") or "").strip().upper()
        if not plate_no:
            continue

        out_value = _cell(row, headers, "out_mileage")
        in_value = _cell(row, headers, "in_mileage")
        distance_value = _cell(row, headers, "distance")
        if (
            out_value in (None, "")
            and in_value in (None, "")
            and distance_value in (None, "")
        ):
            continue

        vehicle = db.scalar(
            select(Vehicle).where(Vehicle.plate_no == plate_no)
        )
        if vehicle is None:
            continue

        parsed_month = _parse_month(_cell(row, headers, "month"))
        now = datetime.now()
        year, month = parsed_month or (now.year, now.month)
        trip_date = date(year, month, 1)
        out_mileage = _parse_decimal(out_value)
        in_mileage = (
            _parse_decimal(in_value)
            if in_value not in (None, "")
            else None
        )
        distance = _parse_decimal(distance_value)
        if in_mileage is not None:
            distance = in_mileage - out_mileage
        status = "CLOSED" if in_mileage is not None else "OUT"

        duplicate = db.scalar(
            select(MileageRecord).where(
                MileageRecord.vehicle_id == vehicle.id,
                MileageRecord.trip_date == trip_date,
                MileageRecord.out_mileage == out_mileage,
                MileageRecord.status == status,
            )
        )
        if duplicate:
            continue

        abnormal = out_mileage < vehicle.current_mileage
        reason = None
        if abnormal:
            reason = "出车里程小于车辆当前里程"
        if in_mileage is not None and in_mileage < out_mileage:
            abnormal = True
            reason = "收车里程小于出车里程"

        db.add(
            MileageRecord(
                vehicle_id=vehicle.id,
                trip_date=trip_date,
                out_mileage=out_mileage,
                in_mileage=in_mileage,
                distance=distance,
                driver_name=str(
                    _cell(row, headers, "driver_name") or ""
                ).strip()
                or None,
                status=status,
                abnormal=abnormal,
                abnormal_reason=reason,
                remark=str(
                    _cell(row, headers, "remark") or ""
                ).strip()
                or None,
                created_by=user.id,
            )
        )
        created += 1

    return created


def _import_maintenances(
    db: Session, workbook, user: User
) -> int:
    ws = _find_sheet(workbook, ["维保"])
    if ws is None:
        return 0
    header_row, headers = _locate_header(ws)
    if "车牌号" not in headers:
        return 0

    created = 0
    today = date.today()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        plate_no = str(_cell(row, headers, "plate_no") or "").strip().upper()
        if not plate_no:
            continue

        raw_date = _cell(row, headers, "maintenance_date")
        maintenance_date = _parse_date(raw_date)
        current_mileage = _parse_decimal(
            _cell(row, headers, "current_mileage")
        )
        maintenance_type = _parse_maintenance_type(
            _cell(row, headers, "maintenance_type")
        )
        items = str(_cell(row, headers, "items") or "").strip()
        amount = _parse_decimal(_cell(row, headers, "amount"))
        if (
            maintenance_date is None
            and not items
            and amount == 0
            and current_mileage == 0
        ):
            continue

        if maintenance_date is None:
            maintenance_date = today

        vehicle = db.scalar(
            select(Vehicle).where(Vehicle.plate_no == plate_no)
        )
        if vehicle is None:
            continue

        duplicate = db.scalar(
            select(MaintenanceRecord).where(
                MaintenanceRecord.vehicle_id == vehicle.id,
                MaintenanceRecord.maintenance_date == maintenance_date,
                MaintenanceRecord.items == (items or "未填写"),
                MaintenanceRecord.amount == amount,
            )
        )
        if duplicate:
            continue

        if not current_mileage:
            current_mileage = vehicle.current_mileage or Decimal("0")

        next_mileage = _parse_decimal(
            _cell(row, headers, "next_mileage")
        )

        db.add(
            MaintenanceRecord(
                vehicle_id=vehicle.id,
                maintenance_date=maintenance_date or today,
                current_mileage=current_mileage,
                maintenance_type=maintenance_type,
                items=items or "未填写",
                amount=amount,
                service_provider=str(
                    _cell(row, headers, "service_provider") or ""
                ).strip()
                or None,
                operator_name=str(
                    _cell(row, headers, "operator_name") or ""
                ).strip()
                or None,
                next_mileage=next_mileage or None,
                remark=str(
                    _cell(row, headers, "remark") or ""
                ).strip()
                or None,
            )
        )
        created += 1

    return created


def _import_reimbursements(
    db: Session, workbook, user: User
) -> int:
    ws = _find_sheet(workbook, ["报销"])
    if ws is None:
        return 0
    header_row, headers = _locate_header(ws)
    if "车牌号" not in headers:
        return 0

    created = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        plate_no = str(_cell(row, headers, "plate_no") or "").strip().upper()
        if not plate_no:
            continue

        fuel = _parse_decimal(_cell(row, headers, "fuel"))
        maintenance = _parse_decimal(_cell(row, headers, "maintenance"))
        toll = _parse_decimal(_cell(row, headers, "toll"))
        parking = _parse_decimal(_cell(row, headers, "parking"))
        other = _parse_decimal(_cell(row, headers, "other"))
        if (
            fuel == 0
            and maintenance == 0
            and toll == 0
            and parking == 0
            and other == 0
        ):
            continue

        vehicle = db.scalar(
            select(Vehicle).where(Vehicle.plate_no == plate_no)
        )
        if vehicle is None:
            continue

        parsed_month = _parse_month(_cell(row, headers, "month"))
        if parsed_month is None:
            now = datetime.now()
            parsed_month = (now.year, now.month)
        year, month = parsed_month
        reimbursement_month = f"{year:04d}-{month:02d}"
        first_day = date(year, month, 1)
        related_mileage = _parse_decimal(
            _cell(row, headers, "related_mileage")
        )

        expense_items = [
            ("FUEL", fuel),
            ("MAINTENANCE", maintenance),
            ("TOLL", toll),
            ("PARKING", parking),
            ("OTHER", other),
        ]
        details = [
            {
                "expense_type": expense_type,
                "expense_date": first_day,
                "amount": amount,
                "related_mileage": (
                    related_mileage if related_mileage else None
                ),
                "invoice_no": None,
                "description": None,
                "attachment_url": None,
            }
            for expense_type, amount in expense_items
            if amount > 0
        ]
        if not details:
            continue

        total = sum(
            (detail["amount"] for detail in details), Decimal("0")
        )
        provided_total = _parse_decimal(_cell(row, headers, "total"))
        if provided_total > total:
            total = provided_total

        applicant_name = str(
            _cell(row, headers, "applicant") or ""
        ).strip() or vehicle.vehicle_manager or user.real_name

        duplicate = db.scalar(
            select(Reimbursement).where(
                Reimbursement.reimbursement_month == reimbursement_month,
                Reimbursement.vehicle_id == vehicle.id,
                Reimbursement.total_amount == total,
            )
        )
        if duplicate:
            continue

        status_text = str(
            _cell(row, headers, "reimbursement_status") or ""
        ).strip()
        if "通过" in status_text or status_text.upper() == "APPROVED":
            status = "APPROVED"
        elif (
            "提交" in status_text
            or "审核" in status_text
            or status_text.upper() == "SUBMITTED"
        ):
            status = "SUBMITTED"
        elif "退回" in status_text or status_text.upper() == "REJECTED":
            status = "REJECTED"
        else:
            status = "DRAFT"

        reimbursement = Reimbursement(
            reimbursement_no=_next_reimbursement_no(
                db, reimbursement_month
            ),
            reimbursement_month=reimbursement_month,
            vehicle_id=vehicle.id,
            project_id=vehicle.project_id,
            applicant_id=user.id,
            applicant_name=applicant_name,
            total_amount=total,
            status=status,
            remark=str(_cell(row, headers, "remark") or "").strip() or None,
        )
        db.add(reimbursement)
        db.flush()

        for detail in details:
            db.add(
                ReimbursementDetail(
                    reimbursement_id=reimbursement.id,
                    **detail,
                )
            )
        created += 1

    return created


def _parse_welder_status(value: Any) -> str:
    text = str(value or "").strip().upper()
    if "在线" in text or text == "ONLINE":
        return "ONLINE"
    if "离线" in text or text == "OFFLINE":
        return "OFFLINE"
    if "故障" in text or text == "FAULT":
        return "FAULT"
    return "ONLINE"


def _import_welders(
    db: Session, workbook
) -> tuple[int, int]:
    ws = _find_sheet(workbook, ["焊机"])
    if ws is None:
        return 0, 0

    header_row, headers = _locate_header(ws, required_key="焊机编号")
    if "焊机编号" not in headers:
        return 0, 0

    created = 0
    updated = 0
    max_welder_id = db.scalar(select(func.max(Welder.id))) or 0
    seen_welder_nos = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        welder_no = str(_cell(row, headers, "welder_no") or "").strip()
        if not welder_no:
            continue
        if welder_no in seen_welder_nos:
            welder_no = f"*{welder_no}"
        if (
            welder_no in seen_welder_nos
            or db.scalar(
                select(Welder).where(Welder.welder_no == welder_no)
            )
        ):
            continue
        seen_welder_nos.add(welder_no)

        welder_code = str(_cell(row, headers, "welder_code") or "").strip()
        project_name = str(_cell(row, headers, "project_name") or "").strip()
        project_manager = str(
            _cell(row, headers, "project_manager") or ""
        ).strip()
        project = _get_or_create_project(
            db, project_name, project_manager
        )

        welder = db.scalar(
            select(Welder).where(Welder.welder_no == welder_no)
        )
        if welder is None:
            if not welder_code:
                max_welder_id += 1
                welder_code = f"HJ-{max_welder_id:06d}"
                while db.scalar(
                    select(Welder).where(
                        Welder.welder_code == welder_code
                    )
                ):
                    max_welder_id += 1
                    welder_code = f"HJ-{max_welder_id:06d}"
            welder = Welder(
                welder_code=welder_code,
                welder_no=welder_no,
                location=str(
                    _cell(row, headers, "welder_location") or ""
                ).strip()
                or None,
                project_id=project.id if project else None,
                welder_manager=str(
                    _cell(row, headers, "welder_manager") or ""
                ).strip()
                or None,
                status=_parse_welder_status(
                    _cell(row, headers, "welder_status")
                ),
                remark=str(_cell(row, headers, "remark") or "").strip()
                or None,
            )
            db.add(welder)
            created += 1
        else:
            if welder_code:
                welder.welder_code = welder_code
            if project:
                welder.project_id = project.id
            location = str(
                _cell(row, headers, "welder_location") or ""
            ).strip()
            if location:
                welder.location = location
            manager = str(
                _cell(row, headers, "welder_manager") or ""
            ).strip()
            if manager:
                welder.welder_manager = manager
            welder.status = _parse_welder_status(
                _cell(row, headers, "welder_status")
            )
            remark = str(_cell(row, headers, "remark") or "").strip()
            if remark:
                welder.remark = remark
            updated += 1

    return created, updated


def import_workbook(
    db: Session, content: bytes, user: User
) -> dict:
    """导入一体化工作簿：车辆、里程、维保和报销。"""
    workbook = load_workbook(io.BytesIO(content), data_only=True)

    created, updated, skipped = _import_vehicles(db, workbook, user)
    mileage_created = _import_mileages(db, workbook, user)
    maintenance_created = _import_maintenances(db, workbook, user)
    reimbursement_created = _import_reimbursements(db, workbook, user)
    welder_created, welder_updated = _import_welders(db, workbook)

    db.flush()

    parts = [f"车辆：新增 {created} 辆，更新 {updated} 辆"]
    if mileage_created:
        parts.append(f"里程：新增 {mileage_created} 条")
    if maintenance_created:
        parts.append(f"维保：新增 {maintenance_created} 条")
    if reimbursement_created:
        parts.append(f"报销：新增 {reimbursement_created} 条")
    if welder_created or welder_updated:
        parts.append(f"焊机：新增 {welder_created} 台，更新 {welder_updated} 台")
    if skipped:
        parts.append(f"跳过空行 {skipped} 条")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "mileage_created": mileage_created,
        "maintenance_created": maintenance_created,
        "reimbursement_created": reimbursement_created,
        "welder_created": welder_created,
        "welder_updated": welder_updated,
        "message": "；".join(parts),
    }


def export_workbook(db: Session) -> io.BytesIO:
    """导出完整台账（对外入口）。"""
    return ExcelService.export_full_workbook(db)


def export_vehicle_workbook(db: Session) -> io.BytesIO:
    """导出车辆台账：车辆及相关业务记录。"""
    sections = [
        ("车辆信息", ExcelService._get_vehicles_data),
        ("里程记录", ExcelService._get_mileages_data),
        ("维保记录", ExcelService._get_maintenances_data),
        ("违章记录", ExcelService._get_violations_data),
        ("油费记录", ExcelService._get_fuels_data),
        ("报销单", ExcelService._get_reimbursements_data),
    ]
    wb = Workbook()
    for index, (title, getter) in enumerate(sections):
        ws = wb.active if index == 0 else wb.create_sheet(title)
        ws.title = title
        data = getter(db)
        headers = list(data[0].keys()) if data else ["无数据"]
        ExcelService._write_table_to_sheet(ws, data, headers, start_row=1)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_welder_workbook(db: Session) -> io.BytesIO:
    """导出焊机台账：焊机档案和焊机巡检。"""
    sections = [
        ("焊机档案", ExcelService._get_welders_data),
        ("焊机巡检", ExcelService._get_welder_inspections_data),
    ]
    wb = Workbook()
    for index, (title, getter) in enumerate(sections):
        ws = wb.active if index == 0 else wb.create_sheet(title)
        ws.title = title
        data = getter(db)
        headers = list(data[0].keys()) if data else ["无数据"]
        ExcelService._write_table_to_sheet(ws, data, headers, start_row=1)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
