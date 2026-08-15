from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from reportlab.graphics.barcode import createBarcodeDrawing
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas


SRC_PDF = Path(r"C:\Users\admin\Desktop\范永林体检报告.pdf")
ROSTER = Path(r"C:\Users\admin\Desktop\1.人员名册.xlsx")
BACKGROUND = Path(r"E:\ruichang-vehicle\tmp\pdfs\cover_base-01.png")
OUT_DIR = Path(r"E:\ruichang-vehicle\output\pdf")

pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


def parse_roster(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = None
    people = []

    for row in sheet.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if headers is None:
            if any(value == "姓名" for value in values):
                headers = values
            continue
        if not values or not values[1]:
            continue
        person = dict(zip(headers, values))
        if person.get("姓名"):
            people.append(person)

    return people


def mask_id(id_text: str) -> str:
    if len(id_text) >= 14:
        return id_text[:6] + "*" * 8 + id_text[-4:]
    return id_text


def gender_from_id(id_text: str) -> str:
    digits = "".join(char for char in id_text if char.isdigit())
    if len(digits) < 17:
        return "未知"
    return "男性" if int(digits[16]) % 2 == 1 else "女性"


def make_cover(person: dict, report_no: str, check_date: str, output_path: Path):
    page_width, page_height = A4

    cover = canvas.Canvas(str(output_path), pagesize=A4)
    cover.drawImage(
        str(BACKGROUND),
        0,
        0,
        width=page_width,
        height=page_height,
    )

    def blank(x0: float, top: float, x1: float, bottom: float):
        cover.setFillColorRGB(1, 1, 1)
        cover.rect(
            x0,
            page_height - bottom,
            x1 - x0,
            bottom - top,
            stroke=0,
            fill=1,
        )

    def draw_text(x: float, top: float, text: str, size: float = 13):
        cover.setFillColorRGB(0, 0, 0)
        cover.setFont("STSong-Light", size)
        cover.drawString(x, page_height - top - size * 0.85, text)

    name = person.get("姓名", "")
    gender = gender_from_id(person.get("身份证号", ""))
    age = person.get("年龄", "")
    phone = person.get("手机号", "")
    id_text = mask_id(person.get("身份证号", ""))
    unit = "华夏高铁运营维护有限公司"
    post = person.get("岗位", "")

    blank(195, 208, 370, 236)
    draw_text(202.5, 215.9, f"名：{name}")

    blank(195, 228, 320, 252)
    draw_text(202.1, 234.1, f"别：{gender}")

    blank(330, 228, 430, 252)
    draw_text(343.0, 233.9, f"龄：{age}岁")

    blank(150, 246, 370, 270)
    draw_text(157.3, 253.0, f"*联系电话：{phone}")

    blank(155, 265, 400, 290)
    draw_text(163.4, 272.1, f"身份证号：{id_text}")

    blank(155, 283, 430, 307)
    draw_text(162.7, 289.6, f"单位名称：{unit}")

    blank(155, 305, 430, 329)
    draw_text(162.3, 311.8, f"工作部门：{post}")

    blank(155, 322, 370, 346)
    draw_text(163.4, 328.0, f"检查日期：{check_date}")

    blank(155, 358, 380, 398)
    draw_text(162.7, 366.5, f"体检号：{report_no}")

    blank(85, 736, 240, 772)
    barcode = createBarcodeDrawing(
        "Code128",
        value=report_no,
        barWidth=0.9,
        barHeight=24,
        humanReadable=False,
    )
    barcode.drawOn(cover, 92, page_height - 768)

    blank(205, 800, 340, 822)
    draw_text(215.7, 806.6, f"{name}·{gender}·{age}岁", size=10)

    blank(315, 800, 400, 822)
    draw_text(327.7, 806.6, report_no, size=10)

    cover.showPage()
    cover.save()


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    people = parse_roster(ROSTER)
    check_date = date.today().isoformat()
    created = []

    for index, person in enumerate(people, start=1):
        report_no = str(26000006 + index)
        name = person.get("姓名", "")
        no = person.get("编号", "")
        filename = f"{no}_{name}_体检报告封面.pdf"
        output_path = OUT_DIR / filename
        make_cover(person, report_no, check_date, output_path)
        created.append(str(output_path))

    print(f"CREATED={len(created)}")
    for path in created:
        print(path)


if __name__ == "__main__":
    main()
